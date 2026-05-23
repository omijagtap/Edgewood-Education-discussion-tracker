import os
import io
import time
import uuid
import threading
import builtins
import openpyxl
from flask import Flask, render_template, request, send_file, jsonify, redirect, url_for
import Disussion_Automate as da
import sheets_helper

app = Flask(__name__)
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'edgewood_secret_key_123')

# Global state for live updates
task_progress = {}
task_results = {}
current_task_id = None

# Intercept print to stream live updates natively
original_print = builtins.print

def custom_print(*args, **kwargs):
    global current_task_id
    text = " ".join(str(a) for a in args)
    if current_task_id and text.strip():
        clean_text = text.strip()
        # Make the prints look cooler and more readable for the UI
        if "Processing thread" in clean_text:
            clean_text = clean_text.replace("Processing thread:", "Analyzing Topic:")
        elif "Fetching" in clean_text:
            clean_text = f"Connecting to Canvas: {clean_text}"
            
        # Ignore empty lines or simple separators
        if not set(clean_text).issubset({'-', '=', ' '}):
            task_progress[current_task_id] = clean_text
            
    original_print(*args, **kwargs)

builtins.print = custom_print

def run_audit(task_id, course_codes_input):
    global current_task_id
    current_task_id = task_id
    task_progress[task_id] = "Initializing Canvas API Connection..."
    
    try:
        if "," in course_codes_input:
            sis_ids = [s.strip() for s in course_codes_input.split(",") if s.strip()]
        else:
            sis_ids = [s.strip() for s in course_codes_input.split() if s.strip()]

        if not sis_ids:
            task_progress[task_id] = "ERROR: No valid SIS IDs found in input."
            return

        if not da.check_connection():
            task_progress[task_id] = "ERROR: API connection failed. Check your token."
            return

        wb = openpyxl.Workbook()
        processed_count = 0

        for sis_id in sis_ids:
            # Business Tracker specific filter
            if not sis_id.upper().startswith("BUS"):
                task_progress[task_id] = f"Skipping {sis_id} (Not a Business course)."
                time.sleep(1)
                continue

            task_progress[task_id] = f"Locating Course: {sis_id}..."
            course = da.find_course(sis_id)
            if not course:
                task_progress[task_id] = f"Could not find Course: {sis_id}. Skipping..."
                time.sleep(1)
                continue
                
            course_id = course["id"]
            course_name = course.get("name", "Unknown Course")
            
            # This calls the original function which has all the print statements we intercept!
            flat_rows = da.collect_consolidated_rows(course_id, sis_id, course_name)
            
            has_real_data = any(r.get("Learner Name") != "(No student posts yet)" for r in flat_rows)

            if not flat_rows or not has_real_data:
                task_progress[task_id] = f"No discussions found in {sis_id}. Skipping..."
                time.sleep(1)
                continue

            task_progress[task_id] = "Compiling Data into Excel Structure..."
            da.build_excel_sheet(wb, flat_rows, sis_id, course_name)
            try:
                task_progress[task_id] = "Syncing Report to Google Sheet Database..."
                sheets_helper.push_to_google_sheet(flat_rows, sis_id, course_name)
                # Bust the dashboard cache so next fetch reflects new audit data
                sheets_helper._dashboard_cache["data"] = None
            except Exception as e:
                original_print(f"Error pushing to Google Sheets: {e}")
            processed_count += 1


        if processed_count == 0:
            task_progress[task_id] = "ERROR: No student discussions found in any course."
            return

        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        task_progress[task_id] = "Finalizing Output File..."
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        task_results[task_id] = {
            "file": output,
            "filename": f"MultiCourse_Discussion_Audit_{sis_ids[0].replace(' ', '_')}.xlsx"
        }
        
        time.sleep(0.5) # Slight pause to let UI catch up
        task_progress[task_id] = "COMPLETE"
    except Exception as e:
        task_progress[task_id] = f"ERROR: {str(e)}"
    finally:
        current_task_id = None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/start', methods=['POST'])
def start_audit():
    data = request.json
    course_codes_input = data.get('course_codes', '')
    if not course_codes_input:
        return jsonify({"error": "Please enter at least one course code."}), 400
        
    task_id = str(uuid.uuid4())
    task_progress[task_id] = "Warming up Audit Engine..."
    
    # Run the scraping in a background thread
    thread = threading.Thread(target=run_audit, args=(task_id, course_codes_input))
    thread.start()
    
    return jsonify({"task_id": task_id})

@app.route('/status/<task_id>')
def status(task_id):
    state = task_progress.get(task_id, "Unknown Task")
    return jsonify({"status": state})

@app.route('/download/<task_id>')
def download(task_id):
    if task_id in task_results:
        res = task_results[task_id]
        return send_file(
            res["file"],
            as_attachment=True,
            download_name=res["filename"],
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    return "File not found or expired.", 404

@app.route('/api/dashboard')
def api_dashboard():
    try:
        data = sheets_helper.get_dashboard_data()
        return jsonify(data)
    except Exception as e:
        original_print(f"Dashboard data fetch failed: {e}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    port = int(os.getenv("PORT", 5000))
    app.run(debug=True, host='0.0.0.0', port=port, threaded=True)
