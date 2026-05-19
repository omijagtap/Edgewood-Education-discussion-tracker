"""
======================================================================
 Edgewood Canvas - Discussion Forum Report (Consolidated Layout)
======================================================================
 Usage: python Disussion_Automate.py
        OR: python Disussion_Automate.py "EDU-862-UED4--Spring-2026"

 Output: A consolidated Excel file with the exact requested format:
         Cohort | Course Name | Topic | Learner Name | Created On | Replied | Replied On | Replied By Name | Replied By Email
======================================================================
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import os
import re
import requests
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[!] openpyxl not found. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Load environment variables
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    import subprocess
    import sys
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "python-dotenv"])
        from dotenv import load_dotenv
        load_dotenv()
    except Exception:
        pass

# ──────────────────────────────────────────────────────────────────
#  CONFIG
# ──────────────────────────────────────────────────────────────────
API_URL      = os.getenv("CANVAS_API_URL", "https://edgewood.instructure.com/api/v1")
ACCESS_TOKEN = os.getenv("CANVAS_ACCESS_TOKEN")
HEADERS      = {"Authorization": f"Bearer {ACCESS_TOKEN}" if ACCESS_TOKEN else ""}
# [PERF] Persistent HTTP session — reuses TCP connections across all API calls
SESSION      = requests.Session()
SESSION.headers.update(HEADERS)
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))

# Define TA Email addresses to check against (case-insensitive)
TA_EMAILS = [
    "asharma@edgewood.edu",
    "seegupta@edgewood.edu",
    "nsukhani@edgewood.edu",
    "ppriyadarshani@edgewood.edu",
    "ssahai@edgewood.edu",
    "vpagaria@edgewood.edu",
    "ssaluja@edgewood.edu",
    "MMehndiratta@edgewood.edu",
    "kkaur@edgewood.edu",
    "pnema@edgewood.edu",
    "aghosh@edgewood.edu",
    "PElavarasu@edgewood.edu",
    "mnaruka@edgewood.edu",
    "mkaur@edgewood.edu",
    "Mmagry@edgewood.edu",
    "RUdhani@edgewood.edu",
    "SobJose@edgewood.edu",
    "Jpatole@edgewood.edu",
    "RMishra@edgewood.edu",
]
# [PERF] Pre-compute as lowercase frozenset for O(1) lookups instead of O(n) per check
TA_EMAILS_SET = frozenset(e.strip().lower() for e in TA_EMAILS if e)

# ──────────────────────────────────────────────────────────────────
#  CANVAS API HELPERS
# ──────────────────────────────────────────────────────────────────

def paginated_get(url, params=None):
    results = []
    if params is None:
        params = {}
    params.setdefault("per_page", 100)
    while url:
        try:
            resp = SESSION.get(url, params=params)  # [PERF] uses persistent session
            if resp.status_code != 200:
                # Log error for better RCA
                if resp.status_code == 401:
                    print(f"  [!] API Unauthorized (401). Check your ACCESS_TOKEN.")
                break
            data = resp.json()
            if isinstance(data, list):
                results.extend(data)
            elif isinstance(data, dict):
                results.append(data)
            url    = resp.links.get("next", {}).get("url")
            params = {}
        except Exception as e:
            print(f"  [!] API Request failed: {e}")
            break
    return results


def check_connection():
    """Verify if the token is valid and who it belongs to."""
    print("  Verifying API connection...")
    url = f"{API_URL}/users/self/profile"
    try:
        resp = SESSION.get(url)  # [PERF] uses persistent session
        if resp.status_code == 200:
            user = resp.json()
            print(f"  [v] Connected as: {user.get('name')} ({user.get('login_id')})")
            return True
        else:
            print(f"  [!] Connection failed (Status {resp.status_code})")
            print(f"      Response: {resp.text}")
            return False
    except Exception as e:
        print(f"  [!] Connection error: {e}")
        return False


def find_course(sis_course_id):
    """Enhanced search to find courses by SIS ID, Course Code, or Name."""
    target_clean = str(sis_course_id).strip()
    
    # 1. Try direct SIS lookup with variations
    # We try: Exact, Dash variations (-- vs -), and Space vs Dash variations
    variations = {target_clean}
    
    # Handle dash variations
    if "-Spring" in target_clean and "--Spring" not in target_clean:
        variations.add(target_clean.replace("-Spring", "--Spring"))
    elif "--Spring" in target_clean:
        variations.add(target_clean.replace("--Spring", "-Spring"))
    
    # Handle space variations (e.g. "Spring 2026" -> "Spring-2026")
    if " " in target_clean:
        variations.add(target_clean.replace(" ", "-"))
    if "-" in target_clean:
        variations.add(target_clean.replace("-", " "))

    for var in variations:
        url  = f"{API_URL}/courses/sis_course_id:{var}"
        resp = SESSION.get(url)  # [PERF] uses persistent session
        if resp.status_code == 200:
            return resp.json()

    # 2. Broad search across visible courses (expanded states)
    print(f"  Direct lookup failed. Searching all courses for '{target_clean}'...")
    search_params = {
        "search_term": target_clean,
        "state[]": ["available", "completed", "claimed", "concluded", "unpublished"]
    }
    courses = paginated_get(f"{API_URL}/courses", params=search_params)
    if courses:
        return courses[0]

    # 3. Manual Fuzzy Match + Diagnostics
    all_courses = paginated_get(f"{API_URL}/courses", params={"state[]": ["available", "completed", "claimed", "concluded", "unpublished"]})
    
    if not all_courses:
        print("  [!] Your token cannot see ANY active or concluded courses in this account.")
        return None

    for c in all_courses:
        c_sis  = (c.get("sis_course_id", "") or "").lower()
        c_code = (c.get("course_code", "") or "").lower()
        c_name = (c.get("name", "") or "").lower()
        t_low  = target_clean.lower()
        
        # Check for partial matches or matches with different separators
        if t_low in c_sis or t_low in c_code or t_low in c_name \
           or c_sis in t_low or c_code in t_low or t_low.replace(" ", "-") in c_sis:
            return c
            
    print(f"  [!] No match found for '{target_clean}'.")
    return None


def get_course_users_with_emails(course_id):
    """Fetches all users in the course to get names, emails, and student IDs."""
    print("  Retrieving course directory (user emails & IDs)...")
    url   = f"{API_URL}/courses/{course_id}/users"
    users = paginated_get(url, params={"include[]": ["email", "sis_user_id", "enrollments"], "per_page": 100})
    user_map = {}
    staff_ids = set()  # [PERF] Collect staff IDs in same pass — eliminates a separate API call
    for u in users:
        uid = u.get("id")
        if uid:
            user_map[uid] = {
                "name"      : u.get("name", "Unknown"),
                "email"     : u.get("email") or u.get("login_id") or "",
                "student_id": u.get("sis_user_id") or u.get("login_id") or "N/A"
            }
            # [PERF] Check enrollments to identify staff — avoids separate get_course_staff_ids call
            enrollments = u.get("enrollments", [])
            for enr in enrollments:
                e_type = enr.get("type", "").lower()
                if e_type in ("teacherenrollment", "taenrollment", "designerenrollment"):
                    staff_ids.add(uid)
                    break
    return user_map, staff_ids


def get_full_view(course_id, topic_id, is_group=False):
    """Fetches topic participants and the entry view (replies)."""
    # The view endpoint for course-level and group-level discussions is usually consistent
    url  = f"{API_URL}/courses/{course_id}/discussion_topics/{topic_id}/view"
    resp = SESSION.get(url)  # [PERF] uses persistent session
    if resp.status_code == 200:
        data         = resp.json()
        participants = {p["id"]: {
                            "name" : p.get("display_name", "Unknown"),
                            "email": p.get("email", "")
                         } for p in data.get("participants", [])}
        return participants, data.get("view", [])
    return {}, []


# ──────────────────────────────────────────────────────────────────
#  DATE FORMATTERS
# ──────────────────────────────────────────────────────────────────

def format_created_on(iso_str):
    """Format ISO timestamp to: '4/8/25, 10:54 PM'."""
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace('Z', '+00:00'))
        dt_local = dt.astimezone()
        m = str(dt_local.month)
        d = str(dt_local.day)
        y = dt_local.strftime("%y")
        hour_12 = dt_local.hour % 12
        if hour_12 == 0:
            hour_12 = 12
        h = str(hour_12)
        minute = dt_local.strftime("%M")
        period = dt_local.strftime("%p")
        return f"{m}/{d}/{y}, {h}:{minute} {period}"
    except Exception:
        return iso_str


def format_replied_on(iso_str):
    """Format ISO timestamp to: 'Apr 24, 2025 10:32 PM'."""
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace('Z', '+00:00'))
        dt_local = dt.astimezone()
        b = dt_local.strftime("%b")
        d = str(dt_local.day)
        year = dt_local.strftime("%Y")
        hour_12 = dt_local.hour % 12
        if hour_12 == 0:
            hour_12 = 12
        h = str(hour_12)
        minute = dt_local.strftime("%M")
        period = dt_local.strftime("%p")
        return f"{b} {d}, {year} {h}:{minute} {period}"
    except Exception:
        return iso_str


def calculate_duration_hours(start_iso, end_iso):
    """Returns difference in decimal hours."""
    if not start_iso or not end_iso:
        return None
    try:
        dt1 = datetime.fromisoformat(start_iso.replace('Z', '+00:00'))
        dt2 = datetime.fromisoformat(end_iso.replace('Z', '+00:00'))
        diff = dt2 - dt1
        return round(diff.total_seconds() / 3600, 2)
    except Exception:
        return None


# ──────────────────────────────────────────────────────────────────
#  DATA PROCESSING
# ──────────────────────────────────────────────────────────────────

def get_cohort_and_course(sis_id, course_name_api):
    """Extract Cohort (e.g. UED5, UC5) and Course (e.g. EDU 817, BUS 910) from SIS ID or Course Name."""
    # Cohort Extraction (Supports UED4, UED5, UC5, UC6 etc.)
    cohort_match = re.search(r"(UED|UC)\d+", sis_id, re.IGNORECASE)
    cohort       = cohort_match.group(0).upper() if cohort_match else "N/A"

    # Course Code Extraction (e.g. EDU 817, EDU 862)
    course_match = re.search(r"^([a-zA-Z]+)[- ]?([0-9]+)", sis_id)
    if course_match:
        course_code = f"{course_match.group(1).upper()} {course_match.group(2)}"
    else:
        # Fallback to cleaning the api course name
        clean_match = re.search(r"^([a-zA-Z]+)\s*([0-9]+)", course_name_api)
        if clean_match:
            course_code = f"{clean_match.group(1).upper()} {clean_match.group(2)}"
        else:
            course_code = course_name_api

    return cohort, course_code


def collect_consolidated_rows(course_id, sis_id, course_name_api):
    cohort, course_code = get_cohort_and_course(sis_id, course_name_api)

    # [PERF] Single API call extracts both user_map AND staff_ids — saves one full paginated request
    user_map, staff_ids = get_course_users_with_emails(course_id)

    print("  Fetching modules...")
    modules = paginated_get(f"{API_URL}/courses/{course_id}/modules")
    print(f"    {len(modules)} module(s) found.")

    print("  Fetching discussion topics...")
    all_topics   = paginated_get(f"{API_URL}/courses/{course_id}/discussion_topics")
    topic_lookup = {t["id"]: t for t in all_topics}
    print(f"    {len(all_topics)} discussion topic(s) found.\n")

    # Group all discussion topics into a queue of (module_name, topic) tuples
    topics_queue = []
    processed_topic_ids = set()

    # [PERF] Fetch all module items in parallel using ThreadPoolExecutor
    def _fetch_mod_items(mod):
        mod_name = mod.get("name", f"Module {mod['id']}")
        items = paginated_get(f"{API_URL}/courses/{course_id}/modules/{mod['id']}/items")
        return mod_name, items

    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = {executor.submit(_fetch_mod_items, m): m for m in modules}
        for future in as_completed(futures):
            mod_name, mod_items = future.result()
            disc_topics = [
                topic_lookup[item["content_id"]]
                for item in mod_items
                if item.get("type") == "Discussion"
                and item.get("content_id") in topic_lookup
            ]
            for topic in disc_topics:
                topics_queue.append((mod_name, topic))
                processed_topic_ids.add(topic["id"])

    # Collect independent/orphan discussion topics that are not in any module
    unprocessed_topics = [t for t in all_topics if t["id"] not in processed_topic_ids]
    for topic in unprocessed_topics:
        # Avoid processing Announcements (which Canvas treats as discussion topics with a flag)
        if topic.get("type") == "Announcement" or topic.get("is_announcement"):
            continue
        topics_queue.append(("General / Independent Discussions", topic))

    flat_rows = []

    # Process each topic in our consolidated queue
    for mod_name, topic in topics_queue:
        topic_title = topic.get("title", "Untitled Discussion")
        print(f"  Processing thread: [{mod_name}] >> {topic_title}")

        # Resolve topic author
        topic_author_id = topic.get("user_id") or (topic.get("author") or {}).get("id")
        topic_author_info = user_map.get(topic_author_id, {}) if topic_author_id else {}
        topic_author_name = topic_author_info.get("name") or (topic.get("author") or {}).get("display_name") or f"User#{topic_author_id or 'Unknown'}"
        topic_author_email = (topic_author_info.get("email") or "").strip().lower()

        # Check if topic was created by a student/learner
        is_student_created = False
        if topic_author_id and topic_author_id not in staff_ids:
            is_student_created = True

        participants, view = get_full_view(course_id, topic["id"])

        # ────────────────────────────────────────────────────────
        # GROUP DISCUSSION SUPPORT
        # (In BUS courses, posts are often hidden inside Groups)
        # ────────────────────────────────────────────────────────
        if topic.get("group_category_id") and not view:
            # If this is a group discussion and the master topic is empty, search sub-groups
            group_topics_url = f"{API_URL}/courses/{course_id}/discussion_topics/{topic['id']}/group_topics"
            group_topics = paginated_get(group_topics_url)
            for gt in group_topics:
                _, g_view = get_full_view(course_id, gt["id"], is_group=True)
                if g_view:
                    view.extend(g_view)

        if is_student_created:
            # ────────────────────────────────────────────────────────
            # MODE A: Student-created Discussion Topic (e.g., Assignment Due Dates)
            # ────────────────────────────────────────────────────────
            # The entire topic itself is the learner's query.
            # Replies inside 'view' are responses.
            created_on = format_created_on(topic.get("created_at"))

            ta_replies = []
            all_replies = []

            for r in view:
                rid      = r.get("user_id")
                r_p_info = participants.get(rid, {})
                r_u_info = user_map.get(rid, {})

                r_name  = r_u_info.get("name") or r_p_info.get("name") or f"User#{rid}"
                r_email = (r_u_info.get("email") or r_p_info.get("email") or "").strip().lower()
                r_time  = r.get("created_at", "")

                is_ta = r_email in TA_EMAILS_SET  # [PERF] O(1) frozenset lookup

                reply_details = {
                    "name": r_name,
                    "email": r_u_info.get("email") or r_p_info.get("email") or "",
                    "time": r_time
                }

                if is_ta:
                    ta_replies.append(reply_details)
                all_replies.append(reply_details)

            # Check SLA
            raw_created_on = topic.get("created_at")
            raw_replied_on = None
            first_responder = ""
            if ta_replies:
                sorted_ta = sorted(ta_replies, key=lambda x: x["time"])
                raw_replied_on = sorted_ta[0]["time"]
                first_responder = sorted_ta[0]["name"]
                replied_on = format_replied_on(raw_replied_on)
                replied = "Yes"
            else:
                replied = "No"
                replied_on = ""

            duration = calculate_duration_hours(raw_created_on, raw_replied_on)
            if duration is not None:
                sla_status = "Delayed Response" if duration > 24 else "On Time Response"
            else:
                sla_status = "SLA Not Applicable" if not is_student_created else "Pending Response"

            # Gather all unique repliers for the entire thread
            names = []
            emails = []
            for rep in all_replies:
                if rep["name"] not in names:
                    names.append(rep["name"])
                if rep["email"] and rep["email"] not in emails:
                    emails.append(rep["email"])
            rep_names_str  = ", ".join(names)
            rep_emails_str = ", ".join(emails)

            flat_rows.append({
                "Cohort"            : cohort,
                "Course Name"       : course_code,
                "Topic"             : topic_title,
                "Learner Name"      : topic_author_name,
                "Student ID"        : topic_author_info.get("student_id", "N/A"),
                "Learner Email"     : topic_author_email,
                "Created On"        : created_on,
                "Replied"           : replied,
                "Replied On"        : replied_on,
                "Duration (Hours)"  : duration if duration is not None else "",
                "SLA Status"        : sla_status,
                "First Responder"   : first_responder,
                "Replied By (Name)" : rep_names_str,
                "Replied By (Email)": rep_emails_str
            })

        else:
            # ────────────────────────────────────────────────────────
            # MODE B: Instructor/Staff-created Discussion Topic (e.g., Generative AI Playground)
            # ────────────────────────────────────────────────────────
            # The top-level entries inside 'view' are student queries.
            if not view:
                # Handle the case where the discussion topic is created but has no student posts yet
                flat_rows.append({
                    "Cohort"            : cohort,
                    "Course Name"       : course_code,
                    "Topic"             : topic_title,
                    "Learner Name"      : "(No student posts yet)",
                    "Student ID"        : "N/A",
                    "Learner Email"     : "",
                    "Created On"        : "",
                    "Replied"           : "N/A",
                    "Replied On"        : "",
                    "Duration (Hours)"  : "",
                    "SLA Status"        : "N/A",
                    "First Responder"   : "",
                    "Replied By (Name)" : "",
                    "Replied By (Email)": ""
                })
            else:
                for entry in view:
                    author_id = entry.get("user_id")
                    # IMPORTANT: Skip if the entry itself was posted by staff (not a learner query)
                    if author_id in staff_ids:
                        continue

                    p_info = participants.get(author_id, {})
                    u_info = user_map.get(author_id, {})

                    l_name  = u_info.get("name") or p_info.get("name") or f"User#{author_id}"
                    l_email = (u_info.get("email") or p_info.get("email") or "").strip().lower()
                    l_sid   = u_info.get("student_id", "N/A")
                    
                    raw_created = entry.get("created_at")
                    created_on  = format_created_on(raw_created)

                    # Check replies to this student post
                    raw_replies = entry.get("replies", [])
                    ta_replies = []
                    all_replies = []

                    for rp in raw_replies:
                        rid = rp.get("user_id")
                        ru_info = user_map.get(rid, {})
                        rp_name = ru_info.get("name") or f"User#{rid}"
                        rp_email = (ru_info.get("email") or "").strip().lower()
                        
                        is_ta = rp_email in TA_EMAILS_SET  # [PERF] O(1) frozenset lookup
                        rep_det = {"name": rp_name, "email": rp_email, "time": rp.get("created_at")}
                        if is_ta: ta_replies.append(rep_det)
                        all_replies.append(rep_det)

                    raw_replied = None
                    first_responder = ""
                    if ta_replies:
                        sorted_ta = sorted(ta_replies, key=lambda x: x["time"])
                        raw_replied = sorted_ta[0]["time"]
                        first_responder = sorted_ta[0]["name"]
                        replied_on = format_replied_on(raw_replied)
                        replied = "Yes"
                    else:
                        replied = "No"
                        replied_on = ""

                    duration = calculate_duration_hours(raw_created, raw_replied)
                    if duration is not None:
                        sla_status = "Delayed Response" if duration > 24 else "On Time Response"
                    else:
                        sla_status = "Pending Response"

                    flat_rows.append({
                        "Cohort"            : cohort,
                        "Course Name"       : course_code,
                        "Topic"             : topic_title,
                        "Learner Name"      : l_name,
                        "Student ID"        : l_sid,
                        "Learner Email"     : l_email,
                        "Created On"        : created_on,
                        "Replied"           : replied,
                        "Replied On"        : replied_on,
                        "Duration (Hours)"  : duration if duration is not None else "",
                        "SLA Status"        : sla_status,
                        "First Responder"   : first_responder,
                        "Replied By (Name)" : ", ".join(set(r["name"] for r in all_replies)),
                        "Replied By (Email)": ", ".join(set(r["email"] for r in all_replies if r["email"]))
                    })

    return flat_rows


# ──────────────────────────────────────────────────────────────────
#  EXCEL SHEET BUILDER
# ──────────────────────────────────────────────────────────────────

C_HDR_BG   = "1F3864"   # dark navy
C_HDR_FG   = "FFFFFF"   # white
C_TITLE_BG = "2E75B6"   # slate blue
C_YES_BG   = "E2EFDA"   # soft green
C_YES_FG   = "375623"
C_NO_BG    = "FCE4D6"   # soft pink
C_NO_FG    = "C00000"
C_ALT_BG   = "F9FBFD"   # very light blue alternate
C_BORDER   = "D9D9D9"

# [PERF] Pre-create reusable style objects once at module level
# Avoids creating thousands of identical Font/PatternFill/Border objects per sheet
_SIDE           = Side(border_style="thin", color=C_BORDER)
CACHED_BORDER   = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
CACHED_HDR_FONT = Font(bold=True, color=C_HDR_FG, size=10)
CACHED_HDR_FILL = PatternFill("solid", fgColor=C_HDR_BG)
CACHED_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# [PERF] Cache per-style fill/font combos used repeatedly in data rows
_FILL_CACHE = {}
_FONT_CACHE = {}
_ALIGN_CACHE = {}

def _get_fill(color):
    if color not in _FILL_CACHE:
        _FILL_CACHE[color] = PatternFill("solid", fgColor=color)
    return _FILL_CACHE[color]

def _get_font(bold, color, size=10):
    key = (bold, color, size)
    if key not in _FONT_CACHE:
        _FONT_CACHE[key] = Font(bold=bold, color=color, size=size)
    return _FONT_CACHE[key]

def _get_align(horizontal, vertical="center", wrap=True):
    key = (horizontal, vertical, wrap)
    if key not in _ALIGN_CACHE:
        _ALIGN_CACHE[key] = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)
    return _ALIGN_CACHE[key]


def thin_border():
    return CACHED_BORDER  # [PERF] Return cached singleton


def build_excel_sheet(wb, flat_rows, sis_id, course_name):
    # Sanitize sheet title (Excel doesn't allow: : \ / ? * [ ])
    clean_title = re.sub(r'[:\\\/\?\*\[\]]', '_', sis_id)
    safe_title = clean_title[:31] # Max 31 chars

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1 and wb.active.title == "Sheet":
        ws = wb.active
        ws.title = safe_title
    else:
        ws = wb.create_sheet(title=safe_title)

    # ── Title Block ────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    t = ws["A1"]
    t.value     = f"EDGEWOOD DISCUSSION BOARD AUDIT  |  {course_name} ({sis_id})"
    t.font      = Font(bold=True, size=12, color=C_HDR_FG)
    t.fill      = PatternFill("solid", fgColor=C_TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Column Headers ─────────────────────────────────────────────
    headers = [
        "Cohort", "Course Name", "Topic", "Learner Name", "Student ID", "Learner Email", 
        "Created On", "Replied", "Replied On", "Duration (Hours)", "SLA Status", 
        "First Responder", "Replied By (Name)", "Replied By (Email)"
    ]
    col_widths = [10, 15, 30, 20, 15, 25, 18, 10, 20, 15, 18, 20, 25, 30]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell           = ws.cell(row=2, column=ci, value=h)
        cell.font      = CACHED_HDR_FONT   # [PERF] reuse cached style
        cell.fill      = CACHED_HDR_FILL
        cell.alignment = CACHED_HDR_ALIGN
        cell.border    = CACHED_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[2].height = 24

    # ── Data Row Injection ─────────────────────────────────────────
    if not flat_rows:
        ws.merge_cells("A3:N3")
        empty = ws["A3"]
        empty.value     = "No discussion boards or learner posts found in this course."
        empty.font      = Font(italic=True, color="808080")
        empty.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 24
        last_row = 3
    else:
        for ri, r in enumerate(flat_rows, start=3):
            is_yes   = r["Replied"] == "Yes"
            is_no    = r["Replied"] == "No"
            is_delayed = r["SLA Status"] == "Delayed Response"
            alt_bg   = C_ALT_BG if ri % 2 == 0 else "FFFFFF"

            # Dynamic styling
            replied_bg = C_YES_BG if is_yes else (C_NO_BG if is_no else alt_bg)
            replied_fg = C_YES_FG if is_yes else (C_NO_FG if is_no else "7F8C8D")

            sla_bg = alt_bg
            sla_fg = "000000"
            if is_delayed:
                sla_bg = "FFC7CE"; sla_fg = "9C0006"
            elif r["SLA Status"] == "On Time Response":
                sla_bg = "C6EFCE"; sla_fg = "006100"

            row_cells = [
                (r["Cohort"],             "center", alt_bg, "000000", False),
                (r["Course Name"],        "center", alt_bg, "000000", False),
                (r["Topic"],              "left",   alt_bg, "000000", False),
                (r["Learner Name"],       "left",   alt_bg, "000000", False),
                (r["Student ID"],         "center", alt_bg, "000000", False),
                (r["Learner Email"],      "left",   alt_bg, "000000", False),
                (r["Created On"],         "center", alt_bg, "000000", False),
                (r["Replied"],            "center", replied_bg, replied_fg, True),
                (r["Replied On"],         "center", alt_bg, "000000", False),
                (r["Duration (Hours)"],   "center", alt_bg, "000000", False),
                (r["SLA Status"],         "center", sla_bg, sla_fg, True),
                (r["First Responder"],    "left",   alt_bg, "000000", True),
                (r["Replied By (Name)"],  "left",   alt_bg, "000000", False),
                (r["Replied By (Email)"], "left",   alt_bg, "000000", False)
            ]

            for ci, (val, align, bg, fg, bold) in enumerate(row_cells, start=1):
                cell           = ws.cell(row=ri, column=ci, value=val)
                cell.font      = _get_font(bold, fg)     # [PERF] cached font
                cell.fill      = _get_fill(bg)            # [PERF] cached fill
                cell.alignment = _get_align(align)         # [PERF] cached alignment
                cell.border    = CACHED_BORDER             # [PERF] cached border
            ws.row_dimensions[ri].height = 20
            last_row = ri

    # ── Summary Analysis Table ─────────────────────────────────────
    sum_start = last_row + 3
    
    # Calculate Metrics
    unique_learners = len(set(r["Learner Name"] for r in flat_rows if r["Learner Name"] and r["Learner Name"] != "(No student posts yet)"))
    total_queries   = len([r for r in flat_rows if r["Replied"] != "N/A" and r["Learner Name"] != "(No student posts yet)"])
    total_replied   = len([r for r in flat_rows if r["Replied"] == "Yes"])
    durations       = [r["Duration (Hours)"] for r in flat_rows if isinstance(r["Duration (Hours)"], (int, float))]
    avg_course_time = round(sum(durations)/len(durations), 2) if durations else 0

    # TA Breakdown Calculation
    ta_stats = {}
    for r in flat_rows:
        ta_name = r.get("First Responder")
        if not ta_name: continue
        if ta_name not in ta_stats:
            ta_stats[ta_name] = {"replies": 0, "ontime": 0, "delayed": 0, "durs": []}
        
        ta_stats[ta_name]["replies"] += 1
        if r["SLA Status"] == "On Time Response": ta_stats[ta_name]["ontime"] += 1
        if r["SLA Status"] == "Delayed Response": ta_stats[ta_name]["delayed"] += 1
        if isinstance(r["Duration (Hours)"], (int, float)):
            ta_stats[ta_name]["durs"].append(r["Duration (Hours)"])

    # Table 1: Course Overview
    ws.merge_cells(f"A{sum_start}:D{sum_start}")
    sh1 = ws[f"A{sum_start}"]; sh1.value = "COURSE OVERVIEW"; sh1.font = Font(bold=True, color="FFFFFF"); sh1.fill = PatternFill("solid", fgColor="4472C4"); sh1.alignment = Alignment(horizontal="center")
    sh1.border = thin_border()
    
    overview = [
        ("Total Unique Learners", unique_learners),
        ("Total Queries Found", total_queries),
        ("Total Replied (TAs)", total_replied),
        ("Avg Course Response Time (Hrs)", avg_course_time)
    ]
    for i, (lab, val) in enumerate(overview, start=1):
        c1 = ws.cell(row=sum_start+i, column=1, value=lab); c1.font = Font(bold=True); c1.border = thin_border()
        c2 = ws.cell(row=sum_start+i, column=2, value=val); c2.alignment = Alignment(horizontal="center"); c2.border = thin_border()
        ws.merge_cells(f"B{sum_start+i}:D{sum_start+i}")
        # Apply border to merged cells
        for col in range(2, 5): ws.cell(row=sum_start+i, column=col).border = thin_border()

    # Table 2: TA Performance Breakdown
    ta_start = sum_start + 6
    ws.merge_cells(f"A{ta_start}:G{ta_start}")
    sh2 = ws[f"A{ta_start}"]; sh2.value = "TA PERFORMANCE BREAKDOWN (TREND ANALYSIS)"; sh2.font = Font(bold=True, color="FFFFFF"); sh2.fill = PatternFill("solid", fgColor="C00000"); sh2.alignment = Alignment(horizontal="center")
    sh2.border = thin_border()
    
    ta_hdrs = ["TA Name", "Replies", "On-Time", "Delayed", "Avg Time (Hrs)", "SLA Violation %"]
    for ci, h in enumerate(ta_hdrs, start=1):
        c = ws.cell(row=ta_start+1, column=ci, value=h)
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D9D9D9"); c.alignment = Alignment(horizontal="center"); c.border = thin_border()

    for i, (name, s) in enumerate(ta_stats.items(), start=2):
        avg_t = round(sum(s["durs"])/len(s["durs"]), 2) if s["durs"] else 0
        sla_v = round((s["delayed"]/s["replies"])*100, 1) if s["replies"] > 0 else 0
        
        cells = [
            (name, "left"),
            (s["replies"], "center"),
            (s["ontime"], "center"),
            (s["delayed"], "center"),
            (avg_t, "center"),
            (f"{sla_v}%", "center")
        ]
        for ci, (v, a) in enumerate(cells, start=1):
            c = ws.cell(row=ta_start+i, column=ci, value=v)
            c.alignment = Alignment(horizontal=a); c.border = thin_border()

    ws.freeze_panes = "A3"


# ──────────────────────────────────────────────────────────────────
#  MAIN RUNNER
# ──────────────────────────────────────────────────────────────────

def main():
    print("\n" + "="*65)
    print("  EDGEWOOD CANVAS -- ENHANCED DISCUSSION AUDIT (SLA + MULTI-TAB)")
    print("="*65)

    if not check_connection():
        sys.exit(1)

    if len(sys.argv) > 1:
        raw_input = " ".join(sys.argv[1:])
    else:
        print("\n  Enter Course SIS IDs (separated by space or comma):")
        print("  Example: EDU-862-UED4--Spring-2026, EDU-817-UED5--Spring-2026")
        raw_input = input("\n  > ").strip()

    if not raw_input:
        print("  [!] No SIS ID entered. Exiting.")
        sys.exit(1)

    # Prioritize comma splitting to allow spaces within a single SIS ID (e.g. "Spring 2026")
    if "," in raw_input:
        sis_ids = [s.strip() for s in raw_input.split(",") if s.strip()]
    else:
        # Fallback to whitespace splitting only if no commas are present
        sis_ids = [s.strip() for s in raw_input.split() if s.strip()]
    
    if not sis_ids:
        print("  [!] No valid SIS IDs found.")
        sys.exit(1)

    wb = openpyxl.Workbook()
    processed_count = 0

    for sis_id in sis_ids:
        print(f"\n" + "-"*50)
        print(f"  Processing Course: {sis_id} ...")
        
        course = find_course(sis_id)
        if not course:
            print(f"  [!] Course not found for: '{sis_id}'. Skipping.")
            continue

        course_id   = course["id"]
        course_name = course.get("name", "Unknown Course")
        print(f"  Found --> [{course_id}] {course_name}")

        # Collect data
        flat_rows = collect_consolidated_rows(course_id, sis_id, course_name)

        # Check if we have any actual learner queries (not just placeholders)
        has_real_data = any(r["Learner Name"] != "(No student posts yet)" for r in flat_rows)

        if not flat_rows or not has_real_data:
            print(f"  [!] No student discussions found in this course. Skipping tab.")
            continue

        # Build sheet in shared workbook
        build_excel_sheet(wb, flat_rows, sis_id, course_name)
        processed_count += 1

    if processed_count == 0:
        print("\n  [!] No courses were successfully processed. Excel not saved.")
        sys.exit(1)

    # Final Save
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    # Use first SIS ID in filename for reference
    ref_id = sis_ids[0].replace(" ", "_")
    fname = os.path.join(SCRIPT_DIR, f"MultiCourse_Discussion_Audit_{timestamp}.xlsx")
    
    wb.save(fname)
    print(f"\n" + "="*65)
    print(f"  Excel saved --> {fname}")
    print(f"  Total Courses Processed: {processed_count}")
    print("  Done!\n")

    try:
        os.startfile(fname)
    except Exception:
        pass


if __name__ == "__main__":
    main()